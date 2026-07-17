import os
import sys
import unittest
import openpyxl

# Ensure backend folder is in path so we can import modules
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from export.excel_export import export_to_excel
from export.summary import build_summary_sheet

class TestExportExcel(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.test_output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
        os.makedirs(cls.test_output_dir, exist_ok=True)
        cls.test_output_path = os.path.join(cls.test_output_dir, "test_requirements_export.xlsx")
        
        # Build a small sample of classified_rows (representing flat rows)
        cls.sample_rows = [
            {
                "category": "Functionality",
                "text": "The BMS shall enable cell balancing automatically.",
                "source_doc": "requirements",
                "line_number": 1,
                "matched_rules": ["shall enable", "cell balancing"],
                "score": 2
            },
            {
                "category": "Performance",
                "text": "The pack shall deliver a minimum of 350 km range.",
                "source_doc": "requirements",
                "line_number": 2,
                "matched_rules": ["minimum of", "range"],
                "score": 2
            },
            {
                "category": "Mechanism",
                "text": "Cell voltage monitoring shall be implemented using a dedicated AFE.",
                "source_doc": "platform",
                "line_number": 3,
                "matched_rules": ["voltage monitoring"],
                "score": 1
            },
            {
                "category": "Fault Management",
                "text": "In the event of thermal runaway, isolate the module.",
                "source_doc": "regulatory",
                "line_number": 4,
                "matched_rules": ["thermal runaway", "isolate"],
                "score": 3
            },
            {
                "category": "Performance",
                "text": "Cell operating temperature shall not exceed 60°C.",
                "source_doc": "regulatory",
                "line_number": 10,
                "matched_rules": ["shall not exceed"],
                "score": 2
            },
            {
                "category": "Functionality",
                "text": "User shall be able to view real-time SoC on the dashboard.",
                "source_doc": "requirements",
                "line_number": 8,
                "matched_rules": ["view", "soc"],
                "score": 2
            }
        ]

    def test_summary_calculation(self):
        print("Asserting summary sheet calculations...")
        summary = build_summary_sheet(self.sample_rows)
        
        # Verify counts
        self.assertEqual(summary["total_statements"], 6)
        self.assertEqual(summary["by_category"]["Functionality"], 2)
        self.assertEqual(summary["by_category"]["Mechanism"], 1)
        self.assertEqual(summary["by_category"]["Performance"], 2)
        self.assertEqual(summary["by_category"]["Fault Management"], 1)
        
        self.assertEqual(summary["by_source"]["requirements"], 3)
        self.assertEqual(summary["by_source"]["platform"], 1)
        self.assertEqual(summary["by_source"]["regulatory"], 2)
        print("[PASS] Summary calculations match sample data.")

    def test_export_file_structure(self):
        print("Exporting data and loading excel workbook...")
        path = export_to_excel(self.sample_rows, self.test_output_path)
        self.assertTrue(os.path.exists(path))
        
        wb = openpyxl.load_workbook(path)
        
        # 1. Assert sheet names match exactly
        expected_sheets = ["Summary", "Functionality", "Mechanism", "Performance", "Fault Management"]
        print(f"Asserting sheet names match: {expected_sheets}")
        self.assertEqual(list(wb.sheetnames), expected_sheets)
        print("[PASS] Sheet names match exactly.")
        
        # 2. Check each category sheet's row count and sorting
        print("Asserting category sheet row counts and sorting...")
        
        # Functionality (2 rows + 1 header = 3 rows)
        ws_func = wb["Functionality"]
        self.assertEqual(ws_func.max_row, 3)
        # Verify sorting: line 1 first, line 8 second
        self.assertEqual(ws_func.cell(row=2, column=3).value, 1)
        self.assertEqual(ws_func.cell(row=3, column=3).value, 8)
        
        # Mechanism (1 row + 1 header = 2 rows)
        ws_mech = wb["Mechanism"]
        self.assertEqual(ws_mech.max_row, 2)
        self.assertEqual(ws_mech.cell(row=2, column=3).value, 3)
        
        # Performance (2 rows + 1 header = 3 rows)
        ws_perf = wb["Performance"]
        self.assertEqual(ws_perf.max_row, 3)
        # Verify sorting: line 2 first, line 10 second
        self.assertEqual(ws_perf.cell(row=2, column=3).value, 2)
        self.assertEqual(ws_perf.cell(row=3, column=3).value, 10)
        
        # Fault Management (1 row + 1 header = 2 rows)
        ws_fault = wb["Fault Management"]
        self.assertEqual(ws_fault.max_row, 2)
        self.assertEqual(ws_fault.cell(row=2, column=3).value, 4)
        
        print("[PASS] Category sheets row counts and sorting are correct.")
        
        # 3. Assert header values match exactly
        print("Asserting header row values...")
        expected_headers = ["Statement", "Source Document", "Line Reference", "Matched Rules", "Score"]
        for sheet_name in expected_sheets[1:]: # Skip Summary
            ws = wb[sheet_name]
            headers = [ws.cell(row=1, column=col).value for col in range(1, 6)]
            self.assertEqual(headers, expected_headers)
        print("[PASS] Headers are exactly correct on all category sheets.")
        
        # 4. Check Summary Sheet styling and values
        print("Asserting Summary sheet values...")
        ws_summary = wb["Summary"]
        self.assertEqual(ws_summary.cell(row=1, column=1).value, "Requirements Analysis Summary")
        self.assertEqual(ws_summary.cell(row=3, column=1).value, "Metric / Category / Source")
        self.assertEqual(ws_summary.cell(row=3, column=2).value, "Count")
        self.assertEqual(ws_summary.cell(row=4, column=2).value, 6) # Total unique
        print("[PASS] Summary sheet contains expected labels and values.")

if __name__ == "__main__":
    unittest.main()
