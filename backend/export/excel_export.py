import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .summary import build_summary_sheet, get_flat_rows

def export_to_excel(rows: list[dict], output_path: str) -> str:
    """
    Exports classified requirement rows to a well-formatted Excel workbook (.xlsx).
    Normalizes inputs, groups by category, adds a Summary sheet and 4 category sheets.
    Saves to output_path and returns output_path.
    """
    flat_rows = get_flat_rows(rows)
    
    # Define categories in fixed order
    categories = ["Functionality", "Mechanism", "Performance", "Fault Management"]
    
    # Group rows by category
    categorized_rows = {cat: [] for cat in categories}
    for r in flat_rows:
        cat = r["category"]
        if cat in categorized_rows:
            categorized_rows[cat].append(r)
            
    # Map internal source doc values to human-readable labels
    source_label_map = {
        "requirements": "Customer Requirements",
        "platform": "Platform / Existing Product",
        "regulatory": "Regulatory Standard"
    }
    
    # Initialize Workbook
    wb = Workbook()
    
    # Remove default worksheet
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)
        
    # Styles
    title_font = Font(name="Arial", size=14, bold=True, color="1F497D")
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True)
    data_font = Font(name="Arial", size=10)
    
    # Very light zebra fill for data rows (even indices)
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Helper styling function
    def style_cell(cell, font=None, fill=None, alignment=None, border=None):
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border

    # 1. SUMMARY SHEET
    ws_summary = wb.create_sheet(title="Summary")
    summary_data = build_summary_sheet(rows)
    
    # Title Block
    ws_summary["A1"] = "Requirements Analysis Summary"
    style_cell(ws_summary["A1"], font=title_font)
    
    # Setup Summary rows
    ws_summary.append([]) # Row 2: Spacer
    ws_summary.append(["Metric / Category / Source", "Count"]) # Row 3: Header
    
    summary_rows = [
        ("Total Unique Classified Statements", summary_data["total_statements"]),
        ("", ""), # Row 5: Spacer
        ("Breakdown by Category", ""), # Row 6: Category Header
        ("   Functionality", summary_data["by_category"].get("Functionality", 0)),
        ("   Mechanism", summary_data["by_category"].get("Mechanism", 0)),
        ("   Performance", summary_data["by_category"].get("Performance", 0)),
        ("   Fault Management", summary_data["by_category"].get("Fault Management", 0)),
        ("", ""), # Row 11: Spacer
        ("Breakdown by Source Document", ""), # Row 12: Source Header
        ("   Customer Requirements", summary_data["by_source"].get("requirements", 0)),
        ("   Platform / Existing Product", summary_data["by_source"].get("platform", 0)),
        ("   Regulatory Standard", summary_data["by_source"].get("regulatory", 0)),
    ]
    
    for label, count in summary_rows:
        ws_summary.append([label, count])
        
    # Format Summary Sheet
    ws_summary.column_dimensions["A"].width = 38
    ws_summary.column_dimensions["B"].width = 15
    
    for r_idx in range(3, ws_summary.max_row + 1):
        cell_a = ws_summary.cell(row=r_idx, column=1)
        cell_b = ws_summary.cell(row=r_idx, column=2)
        val_a = cell_a.value
        val_b = cell_b.value
        
        # Check if it's a section header or main table header
        if val_a in ["Metric / Category / Source", "Breakdown by Category", "Breakdown by Source Document"]:
            style_cell(cell_a, font=header_font, fill=header_fill, alignment=Alignment(horizontal="left", vertical="center"), border=thin_border)
            style_cell(cell_b, font=header_font, fill=header_fill, alignment=Alignment(horizontal="right", vertical="center"), border=thin_border)
        elif val_a == "" and val_b == "":
            # Keep spacers plain
            pass
        else:
            style_cell(cell_a, font=data_font, alignment=Alignment(horizontal="left", vertical="center"), border=thin_border)
            style_cell(cell_b, font=data_font, alignment=Alignment(horizontal="right", vertical="center"), border=thin_border)

    # 2. CATEGORY SHEETS
    headers = ["Statement", "Source Document", "Line Reference", "Matched Rules", "Score"]
    
    for cat_name in categories:
        ws = wb.create_sheet(title=cat_name)
        
        # Freeze panes at A2
        ws.freeze_panes = "A2"
        
        # Append Header Row
        ws.append(headers)
        
        # Format Header Row
        for col_idx in range(1, 6):
            cell = ws.cell(row=1, column=col_idx)
            style_cell(cell, font=header_font, fill=header_fill, alignment=Alignment(horizontal="left" if col_idx != 3 and col_idx != 5 else "right", vertical="center"), border=thin_border)
            
        # Sort and Add Data Rows
        sorted_rows = sorted(categorized_rows[cat_name], key=lambda x: (x.get("line_number", 0), x.get("source_doc", "")))
        
        for idx, r in enumerate(sorted_rows, start=2):
            source_raw = r.get("source_doc", "")
            source_nice = source_label_map.get(source_raw, source_raw)
            matched_rules_str = ", ".join(r.get("matched_rules", []))
            
            row_data = [
                r.get("text", ""),
                source_nice,
                r.get("line_number", 0),
                matched_rules_str,
                r.get("score", 0)
            ]
            ws.append(row_data)
            
            # Apply formatting to data row
            # Apply alternating background fill
            row_fill = zebra_fill if idx % 2 == 0 else None
            
            for col_idx in range(1, 6):
                cell = ws.cell(row=idx, column=col_idx)
                
                # Alignments
                if col_idx == 1:
                    # Statement Column wrapped text
                    align = Alignment(wrap_text=True, vertical="top", horizontal="left")
                elif col_idx in [3, 5]:
                    # Numeric columns right aligned
                    align = Alignment(vertical="top", horizontal="right")
                else:
                    align = Alignment(vertical="top", horizontal="left")
                    
                style_cell(cell, font=data_font, fill=row_fill, alignment=align, border=thin_border)
                
        # Set column widths
        ws.column_dimensions["A"].width = 50  # Fixed width for Statement
        
        # Autofit columns B to E
        for col_idx in range(2, 6):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[col_letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            # Add padding and cap at 60
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)
            
    # Save Workbook to output_path
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
